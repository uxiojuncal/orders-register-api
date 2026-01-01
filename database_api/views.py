import json
import os
import urllib.request
import datetime

from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from .models import Order
from django.shortcuts import get_object_or_404
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
from reportlab.lib.colors import white
from io import BytesIO
from reportlab.lib.utils import ImageReader
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

REQUIRED_FIELDS = ['receiver_name', 'address', 'receiver_phone', 'customer_name']

@csrf_exempt
def export_orders_excel(request):
    """Export all orders to Excel file"""
    if request.method != 'GET':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    # Get all orders ordered by date (newest first)
    orders = Order.objects.all().order_by('-date', '-id')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Pedidos"
    
    # Define headers
    headers = [
        'ID', 'Fecha', 'Estado', 'Cliente', 'Tel. Cliente',
        'Destinatario', 'Tel. Destinatario', 'Dirección',
        'Producto', 'Observaciones', 'Firma'
    ]
    
    # Write headers with styling
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Write data
    for row_num, order in enumerate(orders, 2):
        # Format date
        date_str = order.date.strftime('%d/%m/%Y') if order.date else ''
        
        # Format status
        status_display = order.get_status_display() if hasattr(order, 'get_status_display') else order.status
        
        # Format signature
        signature_str = 'Sí' if order.signature else 'No'
        
        # Write row data
        row_data = [
            order.id,
            date_str,
            status_display,
            order.customer_name or '',
            order.customer_phone or '',
            order.receiver_name or '',
            order.receiver_phone or '',
            order.address or '',
            order.product_name or '',
            order.observations or '',
            signature_str
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            
            # Color code by status
            if col_num == 3:  # Status column
                if order.status == Order.Status.PENDING:
                    cell.fill = PatternFill(start_color="FFF4CC", end_color="FFF4CC", fill_type="solid")
                elif order.status == Order.Status.IN_TRANSIT:
                    cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
                elif order.status == Order.Status.DELIVERED:
                    cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                elif order.status == Order.Status.CANCELLED:
                    cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    
    # Auto-adjust column widths
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        
        for row in ws[column_letter]:
            try:
                if len(str(row.value)) > max_length:
                    max_length = len(str(row.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)  # Max width of 50
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Add filters to headers
    ws.auto_filter.ref = ws.dimensions
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=pedidos_export.xlsx'
    
    # Save workbook to response
    wb.save(response)
    
    return response

def health_check(request):
    return JsonResponse({'status': 'ok'}, status=200)

def parse_date(value):
    # Helper function to parse date strings
    if value is None or value == '':
        return None
    if isinstance(value, str):
        try:
            # Try ISO format first (YYYY-MM-DD)
            return datetime.fromisoformat(value).date()
        except ValueError:
            try:
                # Try other common formats
                return datetime.strptime(value, '%Y-%m-%d').date()
            except ValueError:
                return None
    return value

def generate_order_pdf(request, pk):
    # Get order or return 404
    order = get_object_or_404(Order, id=pk)
    
    # Create PDF response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="pedido_{pk}.pdf"'
    
    # Create PDF
    p = canvas.Canvas(response, pagesize=A4)
    width, height = A4
    
    # Header with nice styling
    p.setFont("Helvetica-Bold", 24)
    p.drawString(1*inch, height - 0.8*inch, "INFORME DE PEDIDO")
    
    # Underline
    p.line(1*inch, height - 0.9*inch, width - 1*inch, height - 0.9*inch)
    
    # Pedido number
    p.setFont("Helvetica", 14)
    p.drawString(1*inch, height - 1.2*inch, f"Pedido #{order.id}")
    
    # Order details
    p.setFont("Helvetica", 12)
    y = height - 1.6*inch
    
    # Format date properly
    date_str = order.date.strftime('%d-%m-%Y') if order.date else 'N/A'
    p.drawString(1*inch, y, f"Fecha: {date_str}")
    y -= 0.3*inch
    
    # Status display
    status_display = order.get_status_display() if hasattr(order, 'get_status_display') else order.status
    p.drawString(1*inch, y, f"Estado: {status_display}")
    y -= 0.5*inch
    
    # Customer info
    p.setFont("Helvetica-Bold", 14)
    p.drawString(1*inch, y, "Información del Cliente")
    y -= 0.3*inch
    
    p.setFont("Helvetica", 12)
    p.drawString(1*inch, y, f"Nombre: {order.customer_name or 'N/A'}")
    y -= 0.25*inch
    
    # Format phone number properly
    customer_phone = f"{order.customer_phone}" if order.customer_phone else 'N/A'
    p.drawString(1*inch, y, f"Teléfono: {customer_phone}")
    y -= 0.5*inch
    
    # Receiver info
    p.setFont("Helvetica-Bold", 14)
    p.drawString(1*inch, y, "Información del Destinatario")
    y -= 0.3*inch
    
    p.setFont("Helvetica", 12)
    p.drawString(1*inch, y, f"Nombre: {order.receiver_name}")
    y -= 0.25*inch
    
    # Format phone number properly
    receiver_phone = f"{order.receiver_phone}" if order.receiver_phone else 'N/A'
    p.drawString(1*inch, y, f"Teléfono: {receiver_phone}")
    y -= 0.25*inch
    p.drawString(1*inch, y, f"Dirección: {order.address}")
    y -= 0.5*inch
    
    # Product info
    p.setFont("Helvetica-Bold", 14)
    p.drawString(1*inch, y, "Producto")
    y -= 0.3*inch
    
    p.setFont("Helvetica", 12)
    p.drawString(1*inch, y, f"{order.product_name or 'N/A'}")
    y -= 0.5*inch
    
    # Observations (kept as 'observations' - correct field name)
    if order.observations:
        p.setFont("Helvetica-Bold", 14)
        p.drawString(1*inch, y, "Observaciones")
        y -= 0.3*inch
        
        p.setFont("Helvetica", 12)
        # Handle long text (wrap if needed)
        p.drawString(1*inch, y, order.observations[:100])  # Limit length
        y -= 0.5*inch
    
    # Signature with error handling
    if order.signature:
        p.setFont("Helvetica-Bold", 14)
        p.drawString(1*inch, y, "Firma del Cliente")
        y -= 0.3*inch
        
        try:
            # Handle both local and remote storage
            if hasattr(order.signature, 'path') and os.path.exists(order.signature.path):
                # Local storage
                p.drawImage(order.signature.path, 1*inch, y - 2*inch, width=3*inch, height=1.5*inch)
            else:
                # Remote storage or file not found
                p.setFont("Helvetica-Oblique", 10)
                p.drawString(1*inch, y - 0.5*inch, "[Firma disponible en el sistema]")
        except Exception as e:
            # Handle any image loading errors
            p.setFont("Helvetica-Oblique", 10)
            p.drawString(1*inch, y - 0.5*inch, f"[Error cargando firma: {str(e)}]")
    
    p.showPage()
    p.save()
    
    return response

def parse_phone(value):
    # Helper function to parse phone numbers
    if value is None or value == '':
        return None
    try:
        clean_value = str(value).replace('-', '').replace(' ', '')
        return int(clean_value)
    except (TypeError, ValueError):
        return None

@csrf_exempt
def create_order(request):
    # Check for correct HTTP method
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    # Try to parse JSON payload
    try:
        payload = json.loads(request.body.decode() or '{}')
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    
    # Check for required fields
    for field in REQUIRED_FIELDS:
        if payload[field] is None or payload[field] == '':
            return JsonResponse({'error': f'Required field {field} cannot be empty'}, status=400)
        if field not in payload:
            return JsonResponse({'error': f'Missing required field: {field}'}, status=400)
        
    # Create the order    
    order = Order.objects.create(
        date=payload['date'],
        customer_name=payload['customer_name'],
        customer_phone=parse_phone(payload['customer_phone']),
        receiver_name=payload['receiver_name'],
        receiver_phone=parse_phone(payload['receiver_phone']),
        product_name=payload['product_name'],
        address=payload['address'],
        observations=payload.get('observations', ''),
        status=payload.get('status', Order.Status.PENDING),
        signature=payload.get('signature', None)
    )

    return JsonResponse({'order_id': order.pk}, status=201)

@csrf_exempt
def update_order(request, pk):
    # Check for correct HTTP method
    if request.method not in ('PATCH', 'PUT'):
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    # Try to parse JSON payload
    try:
        payload = json.loads(request.body.decode() or '{}')
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    # Retrieve the order or return 404
    order = get_object_or_404(Order, pk=pk)

    allowed_fields = {
        'date', 'customer_name', 'customer_phone', 'receiver_name', 'receiver_phone',
        'product_name', 'address', 'observations', 'status'
    }

    if payload.get('status') == Order.Status.DELIVERED:
        if not order.signature:
            return JsonResponse({'error': 'Signature is required to mark order as delivered.'}, status=400)
    
    if 'customer_phone' in payload:
        payload['customer_phone'] = parse_phone(payload['customer_phone'])
    if 'receiver_phone' in payload:
        payload['receiver_phone'] = parse_phone(payload['receiver_phone'])
    
    # Parse date field
    if 'date' in payload:
        payload['date'] = parse_date(payload['date'])

    # Update only allowed fields
    for field, value in payload.items():
        if field in allowed_fields:
            setattr(order, field, value)

    order.save()

    data = {
        'id': order.pk,
        'date': order.date.isoformat() if order.date else None,
        'customer_name': order.customer_name,
        'customer_phone': order.customer_phone,
        'receiver_phone': order.receiver_phone,
        'receiver_name': order.receiver_name,
        'product_name': order.product_name,
        'address': order.address,
        'observations': order.observations,
        'status': order.status,
        'signature': order.signature.url if order.signature else None,
    }
    return JsonResponse({'order': data})

def search_orders(request):
    # Check for correct HTTP method
    if request.method != 'GET':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    # Initialize empty queryset
    orders = Order.objects.none()

    # Apply filters based on query parameters
    if any(request.GET.values()):
        orders = Order.objects.all()

        if status := request.GET.get('status'):
            orders = orders.filter(status=status)
        if receiver_name := request.GET.get('receiver_name'):
            orders = orders.filter(receiver_name__icontains=receiver_name)
        if customer_name := request.GET.get('customer_name'):
            orders = orders.filter(customer_name__icontains=customer_name)
        if id := request.GET.get('id'):
            orders = orders.filter(id=id)
        if date := request.GET.get('date'):
            try:
                day, month = date.split('-')
                day = int(day)
                month = int(month)
                orders = orders.filter(date__day=day, date__month=month)
            except ValueError:
                return JsonResponse({'error': 'Formato de fecha inválido. Use DD-MM.'}, status=400)
        
    return JsonResponse({'orders': list(orders.values())})

@csrf_exempt
def upload_signature(request, pk):
    # Check for correct HTTP method
    if request.method not in ('POST', 'PUT', 'PATCH'):
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    # Retrieve the order or return 404
    order = get_object_or_404(Order, pk=pk)

    # Check if a file is provided
    if 'signature' not in request.FILES:
        return JsonResponse({'error': 'No signature file provided'}, status=400)

    # Save the signature file
    order.signature = request.FILES['signature']
    order.save()

    return JsonResponse({'message': 'Signature uploaded successfully'})

@csrf_exempt
def delete_order(request, pk):
    # Check for correct HTTP method
    if request.method != 'DELETE':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    # Retrieve the order or return 404
    order = get_object_or_404(Order, pk=pk)
    
    # Delete the order
    order.delete()
    
    return JsonResponse({'message': 'Order deleted successfully'})